import openpyxl
import pandas as pd
from DraftEmail import *

class ExcelTracker:
    wb = openpyxl.load_workbook(email_tracker_filename)
    email_tracker_filename = r'C:\Users\ukaea001\Documents\PythonPrograms\PLMB\Reminders\Email_listing_tracker.xlsx'
    sheetname = None
    alert = None

    def __init__(self, email_tracker_filename: str, sheetname: str, alert: str =None) -> None:
        self.email_tracker_filename = email_tracker_filename
        self.sheetname = sheetname
        self.alert = alert
        

    def append_to_sheet(self, first_name, last_name, discipline, email):
        if self.sheetname == "Timesheet":
            df = pd.DataFrame([[TimeStamp.now, first_name, last_name, discipline, email, SendEmail.mailCC_contacts, self.sheetname + ' - '+ self.alert]])
        else: 
            df = pd.DataFrame([[TimeStamp.now, first_name, last_name, discipline, email, SendEmail.mailCC_contacts, self.sheetname]])
        writer = pd.ExcelWriter(self.email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
        ExcelTracker.wb.active = ExcelTracker.wb[self.sheetname]
        # get the max rows of non-empty cells
        filled_rows = len([row for row in ExcelTracker.wb.active if not all([cell.value is None for cell in row])])
        # #Convert dataframe to an Xlsxwrite excel object
        df.to_excel(writer, sheet_name= self.sheetname, index=False, header=False, startrow=filled_rows)
        writer.close()
        print('Added data')
        
    # if sheet does not exist
    def create_new_sheet(self):
        ExcelTracker.wb.create_sheet(self.sheetname)
        df_head = pd.DataFrame([['DATETIME', 'SENT TO', 'DISCIPLINE', 'EMAIL ADDRESS: TO', 'EMAIL ADDRESS: CC', 'TYPE']])
        writer = pd.ExcelWriter(self.email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
        df_head.to_excel(writer, sheet_name= self.sheetname, index=False, header=False)
        writer.close()

    def create_new_file(self):
        # create a new file with headings
        ExcelTracker.wb = openpyxl.Workbook()
        ExcelTracker.wb.save(self.email_tracker_filename)
        print('new file created')

    # check if file does exist 
    def add_to_tracker(self, first_name: str, last_name: str, discipline: str, email: str):
        try:
            if self.sheetname in ExcelTracker.wb.sheetnames: 
                # if file exists, new email log is appended to file
                self.append_to_sheet(first_name, last_name, discipline, email)
            else:
                self.create_new_sheet()
                self.append_to_sheet(first_name, last_name, discipline, email)
        except FileNotFoundError:
            print('file not found')
            self.create_new_file()
            self.create_new_sheet()
            self.append_to_sheet(first_name, last_name, discipline, email)