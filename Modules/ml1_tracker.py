import openpyxl
from datetime import datetime

class ML1Tracker:
    ML1tracker_filename = r'C:\Users\ukaea001\Documents\PythonPrograms\PLMB\ML1 Tracker\WSP-GEN-FOR-RAS-TRA-001_V3_.xlsm'
    ml1_sheetname = 'SAT MDL Tracker'

    @classmethod
    def check_overdue_items(cls, start_row, end_row):
        wb = openpyxl.load_workbook(cls.ML1tracker_filename, read_only=True, keep_vba=True, data_only=True)
        ws = wb[cls.ml1_sheetname]
        overdue = 0 
        # if due date is pass todays date, add 1 to overdue counter
        for rowNum in range(start_row, end_row+1):
            due_date_str = str(ws.cell(row=rowNum, column=36).value)
            if not due_date_str.isspace() and due_date_str != "None":
                due_date_object = datetime.strptime(due_date_str, '%Y-%m-%d %H:%M:%S')
                inputted_cell = str(ws.cell(row=rowNum, column=37).value)
                if datetime.now() > due_date_object and inputted_cell.isspace():
                    overdue += 1
        wb.close()
        return overdue