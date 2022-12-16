# Existing file 
import codecs
import win32com.client as win32
from datetime import datetime
import calendar
import os
import pandas as pd
import openpyxl
from xlsxwriter import Workbook

# access signature and add to email body
sig_files_path = ''
sig_html_path = ''
img_path = ''

# 'C:\Users\ukaea001\AppData\Roaming\Microsoft\Signatures\Elizabeth Adejumo.htm'

# Finds the path to outlook signature files with signature name
signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path)
# specifies the name of the HTML version  of the stored signature
html_doc = os.path.join((os.environ['USERPROFILE']), sig_html_path)
html_doc = html_doc.replace('\\\\', '\\') # removes escape backlashes from path string

html_file = codecs.open(html_doc, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
signature_code = html_file.read() #Writes contents of HTML signature file to a string

signature_code = signature_code.replace(('Elizabeth Adejumo_files/'), signature_path) #Replaces local directory with full directory path
html_file.close()

# email tracking list excel file location
email_tracker_filename = ''
invoice_sheetname = 'Invoice'

#INPUTS 
# scheduled time for invoice alert
# invoice_alert = "6 17:16:00"
now = datetime.now()

# dictionary of contact list 
mailTo_contacts = {

}

mailCC_contacts = ''

def send_invoice_email(): 
    # loop to send emails to each contact
    for k, v in mailTo_contacts.items():
        try:
            # create the new email 
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)

            # set email subject
            previous_month = calendar.month_name[now.month-1]
            mail.Subject = f'PLMB {v[2]} - Invoice Reminder for {previous_month} ' + now.strftime('%Y')

            # set receiver email
            mail.To = v[1]
            mail.cc = mailCC_contacts

            # write email content
            email_message = r'''
            Hi {2},<br><br>
            This is a reminder that we are still awaiting your invoice for <b>{0} {1}</b>. <br><br>
            Could you please send this invoice as soon as possible.
            Please ensure that when you send through your invoices, you include a copy of the timesheets to back up the invoice. <br><br>
            Please ignore this email if you have sent your invoice previously. <br><br>
            Thanks.<br><br>
            '''.format(previous_month, now.year, k)

            mail.HTMLBody = email_message + signature_code
            inspector = mail.getInspector
            # Adding signature image
            doc = inspector.WordEditor
            selection = doc.Content
            selection.Find.Text = "insert image"
            selection.Find.Execute()
            selection.Text = ""
            img = selection.InlineShapes.AddPicture(img_path, 0, 1)
            mail.display()
            # mail.Send()
            print('Invoice email alert activated - message sent')
            add_to_mailing_list(v[0], v[1], v[2])
        except Exception as e:
            print("Invoice email alert failed to send: " + str(e))

# create mailing list to track emails
def add_to_mailing_list(name, email, discipline):
    # check if file does exist 
    try:
        # if file exists, new email log is appended to file
        append_to_mailing_list(name, email, discipline)
        print('file found')
    except FileNotFoundError:
        print('file not found')
        # create a new file with headings
        wb = Workbook()
        wb.save(email_tracker_filename= email_tracker_filename)
        print('new file created')
        # create new sheet
        create_new_sheet()
        append_to_mailing_list(name, email, discipline)
    
def append_to_mailing_list(name, email, discipline):
    wb = openpyxl.load_workbook(email_tracker_filename)
    if invoice_sheetname in wb.sheetnames: 
        # dataframe Datetime	Sent to	Email address
        df = pd.DataFrame([[now, name, discipline, email, mailCC_contacts, invoice_sheetname]])
        writer = pd.ExcelWriter(email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
        wb.active = wb[invoice_sheetname]
        # get the max rows of non-empty cells
        filled_rows = len([row for row in wb.active if not all([cell.value is None for cell in row])])
        # #Convert dataframe to an Xlsxwrite excel object
        df.to_excel(writer, sheet_name= invoice_sheetname, index=False, header=False, startrow=filled_rows)
        writer.close()
        print('Added data')
    else:
        create_new_sheet()
        append_to_mailing_list(name, email, discipline)

def create_new_sheet():
    wb = openpyxl.load_workbook(email_tracker_filename)
    # if sheet does not exist
    wb.create_sheet(invoice_sheetname)
    df_head = pd.DataFrame([['DATETIME', 'SENT TO', 'DISCIPLINE', 'EMAIL ADDRESS: TO', 'EMAIL ADDRESS: CC', 'TYPE']])
    writer = pd.ExcelWriter(email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
    df_head.to_excel(writer, sheet_name= invoice_sheetname, index=False, header=False)
    writer.close()
    print('new sheet created')


send_invoice_email()


