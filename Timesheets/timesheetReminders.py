import codecs
import win32com.client as win32
from datetime import datetime, timedelta
import calendar
import os
import pandas as pd
import openpyxl

# access signature and add to email body
sig_files_path = 'AppData\Roaming\Microsoft\Signatures\Elizabeth Adejumo_py_files\\'
sig_html_path = 'AppData\Roaming\Microsoft\Signatures\Elizabeth Adejumo_py.htm'
img_path = r'C:\Users\ukaea001\AppData\Roaming\Microsoft\Signatures\Elizabeth Adejumo_files\image001.png' # needed

# 'C:\Users\ukaea001\AppData\Roaming\Microsoft\Signatures\Elizabeth Adejumo.htm'

# Finds the path to outlook signature files with signature name
signature_path = os.path.join((os.environ['USERPROFILE']), sig_files_path)
# specifies the name of the HTML version  of the stored signature
html_doc = os.path.join((os.environ['USERPROFILE']), sig_html_path)
html_doc = html_doc.replace('\\\\', '\\') # removes escape backlashes from path string

html_file = codecs.open(html_doc, 'r', 'utf-8', errors='ignore') #Opens HTML file and converts to UTF-8, ignoring errors
signature_code = html_file.read() #Writes contents of HTML signature file to a string

signature_code = signature_code.replace(('Elizabeth Adejumo_files/'), signature_path) #Replaces local directory with full directory path # needed
html_file.close()

# email tracking list excel file location
email_tracker_filename = r'C:\Users\ukaea001\Documents\PythonPrograms\PLMB\Reminders\Email_listing_tracker.xlsx'
timesheet_sheetname = 'Timesheet'

#INPUTS 
# scheduled time for timesheet alert
# timesheet_alert = "6 17:16:00"
now = datetime.now()
today_date = int(now.strftime('%d'))
starting_monday = now - timedelta(4)
week_range = starting_monday.strftime('%d %b %Y') + ' to ' + now.strftime('%d %b %Y')
last_day_of_month = calendar.monthrange(now.year, now.month)[1]

# dictionaries
mailTo_contacts = {
    'David': ['David Starbuck', 'david.starbuck@wsp.com', 'Depot and PDAP'],
    # 'Benny': ['Benny Lee', 'benny.lee@wsp.com', 'AFC HK'],
    # 'Nelson': ['Nelson Ng', 'nelson.ng@wsp.com', 'PDAP HK'],
    # 'Elizabeth': ['Elizabeth Ade', 'elizjumo01@gmail.com', 'TRIAL']
}

mailCC_contacts = 'carolina.morales@wsp.com'

month_end = {
    'mail_subject': 'Month end tomorrow',
    'message': r'''
        This is a reminder that it is month end tomorrow. Could you please send your timesheet for this week by <b>12PM tomorrow</b>. <br><br>
        Thanks.<br><br>
        ''',
    'alert': 'Month end'
}

friday_month_end = {
    'mail_subject': 'Month end',
    'message': r'''
        This is a reminder that it is month end. Could you please send your timesheet for this week by <b>COB today</b>. <br><br>
        Please ignore this email if you have sent your timesheet previously. <br><br>
        Thanks.<br><br>
        ''',
    'alert': 'Month end (during weekend)'
}

friday_alert = {
    'mail_subject': week_range,
    'message': r'''
        This is a reminder to send your timesheet for {0} by <b>COB today</b>. <br><br>
        Please ignore this email if you have sent your timesheet previously. <br><br>
        Thanks.<br><br>
        '''.format(week_range),
    'alert': 'Friday timesheet'
}

def timesheet_alert(mail_subject, message, alert):
    # loop to send emails to each contact
    for k, v in mailTo_contacts.items():
        try:
            # create the new email 
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)

            mail.Subject = f'PLMB {v[2]} Timesheet reminder - ' + mail_subject
            mail.To = v[1]
            mail.cc = mailCC_contacts
            email_message = 'Hi {0},<br><br>'.format(k) + message
            mail.HTMLBody = email_message + signature_code
            inspector = mail.getInspector

            # Adding signature image
            doc = inspector.WordEditor
            selection = doc.Content
            selection.Find.Text = "insert image"
            selection.Find.Execute()
            selection.Text = ""
            img = selection.InlineShapes.AddPicture(img_path, 0, 1)
            mail.display()
            mail.Send()
            print(alert + ' email alert - message sent')
            add_to_mailing_list(v[0], v[1], v[2], alert)
        except Exception as e:
            print("Timesheet email alert failed to send: " + str(e))

# create mailing list to track emails
def add_to_mailing_list(name, email, discipline, alert):
    # check if file does exist 
    try:
        # if file exists, new email log is appended to file
        append_to_mailing_list(name, email, discipline, alert)
        print('file found')
    except FileNotFoundError:
        print('file not found')
        # create a new file with headings
        wb = openpyxl.Workbook()
        wb.save(email_tracker_filename)
        print('new file created')
        # create new sheet
        create_new_sheet()
        append_to_mailing_list(name, email, discipline, alert)
    
def append_to_mailing_list(name, email, discipline, alert):
    wb = openpyxl.load_workbook(email_tracker_filename)
    if timesheet_sheetname in wb.sheetnames: 
        # dataframe Datetime	Sent to	Email address
        df = pd.DataFrame([[now, name, discipline, email, mailCC_contacts, timesheet_sheetname + ' - '+ alert]])
        writer = pd.ExcelWriter(email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
        wb.active = wb[timesheet_sheetname]
        # get the max rows of non-empty cells
        filled_rows = len([row for row in wb.active if not all([cell.value is None for cell in row])])
        # #Convert dataframe to an Xlsxwrite excel object
        df.to_excel(writer, sheet_name= timesheet_sheetname, index=False, header=False, startrow=filled_rows)
        writer.close()
        print('Added data')
    else:
        create_new_sheet()
        append_to_mailing_list(name, email, discipline, alert)

def create_new_sheet():
    wb = openpyxl.load_workbook(email_tracker_filename)
    # if sheet does not exist
    wb.create_sheet(timesheet_sheetname)
    df_head = pd.DataFrame([['DATETIME', 'SENT TO', 'DISCIPLINE', 'EMAIL ADDRESS: TO', 'EMAIL ADDRESS: CC', 'TYPE']])
    writer = pd.ExcelWriter(email_tracker_filename, if_sheet_exists='overlay', mode='a', engine='openpyxl')
    df_head.to_excel(writer, sheet_name= timesheet_sheetname, index=False, header=False)
    writer.close()
    print('new sheet created')

def send_timesheet_email(): 
    if today_date == last_day_of_month - 1:
        timesheet_alert(month_end['mail_subject'], month_end['message'], month_end['alert'])
    elif now.strftime('%A') == 'Friday' and today_date != last_day_of_month:   
        if today_date + 1 == last_day_of_month or today_date + 2 == last_day_of_month or today_date + 3 == last_day_of_month:
            timesheet_alert(friday_month_end['mail_subject'], friday_month_end['message'], friday_month_end['alert']) 
        else:
            timesheet_alert(friday_alert['mail_subject'], friday_alert['message'], friday_alert['alert'])
    else:
        print("no timesheet reminder today")

send_timesheet_email()

